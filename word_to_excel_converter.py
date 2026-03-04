"""
Word -> Excel Converter (Structured)
Author: Ece
Usage:
  python word_to_excel_converter.py --input input.docx --output output.xlsx
"""

from __future__ import annotations

import argparse
from dataclasses import dataclass
from typing import Optional, List

from docx import Document
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


@dataclass
class RowItem:
    idx: int
    level: int
    kind: str       # "HEADING" or "TEXT"
    text: str


def infer_heading_level(paragraph_style_name: str) -> Optional[int]:
    """
    Tries to infer heading level from Word style name like:
    'Heading 1', 'Heading 2', 'Başlık 1', 'Baslik 1', etc.
    """
    if not paragraph_style_name:
        return None

    s = paragraph_style_name.strip().lower()

    # common English
    if "heading" in s:
        parts = s.split()
        for p in parts:
            if p.isdigit():
                return int(p)

    # common Turkish variants
    # "Başlık 1" / "Baslik 1"
    if "başlık" in s or "baslik" in s:
        parts = s.replace("başlık", "baslik").split()
        for p in parts:
            if p.isdigit():
                return int(p)

    return None


def parse_docx(docx_path: str) -> List[RowItem]:
    doc = Document(docx_path)
    items: List[RowItem] = []

    idx = 1
    current_section = ""
    for p in doc.paragraphs:
        text = (p.text or "").strip()
        if not text:
            continue

        style_name = getattr(p.style, "name", "") if p.style else ""
        level = infer_heading_level(style_name)

        if level is not None:
            current_section = text
            items.append(RowItem(idx=idx, level=level, kind="HEADING", text=text))
        else:
            # content row
            # indent level "0" unless we have a section heading earlier
            items.append(RowItem(idx=idx, level=0, kind="TEXT", text=text))

        idx += 1

    return items


def write_xlsx(items: List[RowItem], output_path: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Converted"

    headers = ["No", "Type", "HeadingLevel", "Text"]
    ws.append(headers)

    # header styling
    header_font = Font(bold=True)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # rows
    for it in items:
        ws.append([it.idx, it.kind, it.level, it.text])

        r = ws.max_row
        if it.kind == "HEADING":
            # make headings bold, and slightly larger based on level
            size = 14 if it.level == 1 else 13 if it.level == 2 else 12
            ws.cell(row=r, column=4).font = Font(bold=True, size=size)
            ws.cell(row=r, column=2).font = Font(bold=True)
            ws.cell(row=r, column=3).font = Font(bold=True)
        else:
            ws.cell(row=r, column=4).alignment = Alignment(wrap_text=True, vertical="top")

    # column widths
    col_widths = {1: 6, 2: 10, 3: 14, 4: 90}
    for col, w in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = "A2"
    wb.save(output_path)


def main():
    parser = argparse.ArgumentParser(description="Convert structured Word (.docx) to Excel (.xlsx).")
    parser.add_argument("--input", "-i", required=True, help="Path to input .docx file")
    parser.add_argument("--output", "-o", default="output.xlsx", help="Path to output .xlsx file")
    args = parser.parse_args()

    items = parse_docx(args.input)
    if not items:
        raise SystemExit("No content found in the document.")

    write_xlsx(items, args.output)
    print(f" Done! Wrote {len(items)} rows to: {args.output}")


if __name__ == "__main__":
    main()
